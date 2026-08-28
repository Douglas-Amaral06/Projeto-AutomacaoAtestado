from openpyxl import Workbook, load_workbook

from app.spreadsheet_pipeline import (
    append_received_document,
    find_employee,
    remove_received_document,
    safe_excel_value,
    source_records,
)


def test_safe_excel_value_blocks_formula_injection_without_changing_native_types():
    for payload in ("=1+1", "+cmd", "-2+3", "@SUM(A1:A2)", "  =HYPERLINK(\"x\")", "\t=cmd"):
        sanitized = safe_excel_value(payload)
        assert sanitized == "'" + payload

    assert safe_excel_value("texto normal") == "texto normal"
    assert safe_excel_value(123) == 123
    assert safe_excel_value(None) is None


def test_source_workbook_is_closed_when_generator_is_stopped_early(tmp_path):
    source = tmp_path / "base.xlsx"
    save_workbook(
        source,
        ["CHAPA", "NOME", "CPF"],
        [[1, "Pessoa Um", "52998224725"], [2, "Pessoa Dois", "01234567890"]],
        "ATIVOS",
    )
    records = source_records(source)
    assert next(records)["NOME"] == "Pessoa Um"
    records.close()

    source.unlink()
    assert not source.exists()


def save_workbook(path, headers, rows, title):
    workbook = Workbook()
    sheet = workbook.active
    sheet.title = title
    sheet.append(headers)
    for row in rows:
        sheet.append(row)
    workbook.save(path)


def test_automatic_enrichment_requires_exact_cpf_and_name(tmp_path, monkeypatch):
    source = tmp_path / "base.xlsx"
    save_workbook(
        source,
        ["CHAPA", "NOME", "CPF", "EMAILFUNCIONARIO", "TELEFONEFUNCIONARIO", "NOMETOMADOR"],
        [[321, "Joao Gomes", "01234567890", "joao@empresa.test", "11912345678", "Empresa Segura"]],
        "Base ATIVOS",
    )
    monkeypatch.setenv("PIPELINE_BASE_GERAL_PATH", str(source))

    employee, status = find_employee("João Gome", "012.345.678-90")
    assert status == "ENCONTRADO_CPF_NOME"
    assert employee == {
        "matricula": 321,
        "telefone": "11912345678",
        "email": "joao@empresa.test",
        "empresa": "Empresa Segura",
    }
    assert find_employee("Pessoa Diferente", "012.345.678-90") == (None, "NAO_ENCONTRADO")


def test_received_document_is_appended_once_and_can_be_removed(tmp_path, monkeypatch):
    target = tmp_path / "atestados.xlsx"
    save_workbook(
        target,
        ["MATRÍCULA", "NOME", "TELEFONE", "E-MAIL", "EMPRESA", "TIPO DE DOCUMENTO", "DATA DE RECEBIMENTO", "DATA DO ATESTADO", "CID", "QUANTIDADE DE DIAS", "ENVIADO PARA", "RESPONSÁVEL", "MEIO DE ENVIO"],
        [
            [100, "Registro anterior", "", "", "", "ATESTADO", "", "", "", "", "DP NACIONAL", "Analista anterior", "WhatsApp"],
            ["", "", "", "", "", "", "", "", "", "", "DP NACIONAL", "Nome indevido", "WhatsApp"],
        ],
        "2026",
    )
    monkeypatch.setenv("PIPELINE_ATESTADOS_PATH", str(target))
    extracted = {"nome": "João Gomes", "cpf": "01234567890", "tipo_documento": "atestado", "data_atestado": "2026-08-05", "cid": "Z00", "dias_afastamento": 2}
    employee = {"matricula": 321, "telefone": "11912345678", "email": "joao@empresa.test", "empresa": "Empresa Segura"}

    first = append_received_document(extracted, employee, "hash-unico", "ENCONTRADO_CPF_NOME")
    second = append_received_document(extracted, employee, "hash-unico", "ENCONTRADO_CPF_NOME")
    assert first == {"status": "REGISTRADO", "linha": 3}
    assert second == {"status": "ATUALIZADO", "linha": 3}

    workbook = load_workbook(target)
    sheet = workbook["2026"]
    headers = {cell.value: cell.column for cell in sheet[1] if cell.value}
    assert sheet.cell(3, headers["MATRÍCULA"]).value == 321
    assert sheet.cell(3, headers["NOME"]).value == "João Gomes"
    assert sheet.cell(3, headers["CPF"]).value == "01234567890"
    assert sheet.cell(3, headers["E-MAIL"]).value == "joao@empresa.test"
    assert sheet.cell(3, headers["ENVIADO PARA"]).value is None
    assert sheet.cell(3, headers["RESPONSÁVEL"]).value is None
    assert sheet.cell(3, headers["MEIO DE ENVIO"]).value is None
    assert sheet.column_dimensions[sheet.cell(1, headers["ID TÉCNICO DO ARQUIVO"]).column_letter].hidden
    workbook.close()

    assert remove_received_document("hash-unico") is True
    workbook = load_workbook(target)
    assert workbook["2026"].max_row == 2
    workbook.close()

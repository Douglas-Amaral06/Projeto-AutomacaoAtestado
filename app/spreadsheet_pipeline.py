import os
import re
import threading
import unicodedata
import uuid
import time
from contextlib import contextmanager
from collections import defaultdict
from copy import copy
from datetime import date, timedelta
from difflib import SequenceMatcher
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import PatternFill

from .database import connect
from .security import utc_now


_workbook_write_lock = threading.Lock()
EXCEL_MAX_ROWS = 1_048_576


@contextmanager
def _spreadsheet_lease(timeout_seconds: float = 15.0):
    owner = uuid.uuid4().hex
    deadline = time.monotonic() + timeout_seconds
    acquired = False
    while time.monotonic() < deadline:
        now = utc_now()
        expires = now + timedelta(minutes=30)
        try:
            with connect() as connection:
                connection.execute("BEGIN IMMEDIATE")
                connection.execute(
                    "INSERT OR IGNORE INTO recursos_lock(nome,owner,expires_at) VALUES('planilha',NULL,NULL)"
                )
                cursor = connection.execute(
                    """UPDATE recursos_lock SET owner=?,expires_at=? WHERE nome='planilha'
                       AND (owner IS NULL OR expires_at IS NULL OR expires_at<=?)""",
                    (owner, expires.isoformat(), now.isoformat()),
                )
                acquired = cursor.rowcount == 1
            if acquired:
                break
        except Exception:
            acquired = False
        time.sleep(0.1)
    if not acquired:
        raise RuntimeError("A planilha está ocupada por outro processo. Tente novamente.")
    try:
        yield
    finally:
        with connect() as connection:
            connection.execute(
                "UPDATE recursos_lock SET owner=NULL,expires_at=NULL WHERE nome='planilha' AND owner=?",
                (owner,),
            )


def safe_excel_value(value):
    """Força texto potencialmente executável a permanecer literal no Excel."""
    if isinstance(value, str) and value.lstrip().startswith(("=", "+", "-", "@", "\t", "\r", "\n")):
        return "'" + value
    return value


def normalize(value) -> str:
    text = unicodedata.normalize("NFKD", str(value or "")).encode("ascii", "ignore").decode().upper()
    return re.sub(r"[^A-Z0-9]", "", text)


def digits(value) -> str:
    text = str(value or "").strip()
    if text.endswith(".0"):
        text = text[:-2]
    return re.sub(r"\D", "", text).lstrip("0")


def configured_path(variable: str) -> Path:
    raw = os.getenv(variable, "")
    path = Path(raw).expanduser().resolve() if raw else None
    if not path or not path.is_file() or path.suffix.lower() != ".xlsx":
        raise RuntimeError(f"{variable} nao aponta para um XLSX valido")
    return path


def header_map(sheet) -> dict[str, int]:
    return {normalize(cell.value): index for index, cell in enumerate(sheet[1], 1) if cell.value is not None}


def value(row, index):
    return row[index - 1] if index and index <= len(row) else None


def source_records(path: Path):
    workbook = load_workbook(path, read_only=True, data_only=True)
    try:
        ordered = sorted(workbook.worksheets, key=lambda sheet: ("ATIVOS" not in normalize(sheet.title), "GERAL" not in normalize(sheet.title)))
        for sheet in ordered:
            headers = header_map(sheet)
            if not all(key in headers for key in ("CHAPA", "NOME", "CPF")):
                continue
            fields = {
                key: headers.get(key)
                for key in ("CHAPA", "NOME", "CPF", "EMAILFUNCIONARIO", "TELEFONEFUNCIONARIO", "TELEFONE2", "TELEFONE3", "NOMETOMADOR", "NOMESECAO", "NOMEFILIAL")
            }
            blank_run = 0
            found_data = False
            for row in sheet.iter_rows(min_row=2, values_only=True):
                if not any(value(row, fields[key]) not in (None, "") for key in ("CHAPA", "NOME", "CPF")):
                    blank_run += 1
                    if found_data and blank_run >= 1000:
                        break
                    continue
                blank_run = 0
                found_data = True
                yield {key: value(row, index) for key, index in fields.items()}
    finally:
        workbook.close()


def first_present(record: dict, *keys: str):
    return next((record.get(key) for key in keys if record.get(key) not in (None, "")), None)


def build_indexes(path: Path):
    by_cpf_name = defaultdict(list)
    by_cpf = defaultdict(list)
    by_chapa_name = defaultdict(list)
    by_chapa = defaultdict(list)
    seen = set()
    for record in source_records(path):
        identity = (digits(record["CHAPA"]), digits(record["CPF"]), normalize(record["NOME"]))
        if identity in seen:
            continue
        seen.add(identity)
        if identity[1] and identity[2]:
            by_cpf_name[(identity[1], identity[2])].append(record)
            by_cpf[identity[1]].append(record)
        if identity[0] and identity[2]:
            by_chapa_name[(identity[0], identity[2])].append(record)
        if identity[0]:
            by_chapa[identity[0]].append(record)
    return by_cpf_name, by_cpf, by_chapa_name, by_chapa


def find_employee(name, cpf) -> tuple[dict | None, str]:
    """Localiza uma pessoa unicamente por CPF + nome, sem expor a fonte."""
    if not normalize(name) or not digits(cpf):
        return None, "DADOS_INSUFICIENTES"
    source_path = configured_path("PIPELINE_BASE_GERAL_PATH")
    by_cpf_name, by_cpf, _, _ = build_indexes(source_path)
    matches = by_cpf_name.get((digits(cpf), normalize(name)), [])
    if not matches:
        cpf_matches = by_cpf.get(digits(cpf), [])
        if len(cpf_matches) == 1:
            similarity = SequenceMatcher(None, normalize(name), normalize(cpf_matches[0].get("NOME"))).ratio()
            if similarity >= 0.82:
                matches = cpf_matches
    if len(matches) != 1:
        return None, "NAO_ENCONTRADO" if not matches else "REVISAR_DUPLICIDADE"
    record = matches[0]
    return {
        "matricula": record.get("CHAPA"),
        "telefone": first_present(record, "TELEFONEFUNCIONARIO", "TELEFONE2", "TELEFONE3"),
        "email": record.get("EMAILFUNCIONARIO"),
        "empresa": first_present(record, "NOMETOMADOR", "NOMESECAO", "NOMEFILIAL"),
    }, "ENCONTRADO_CPF_NOME"


def _last_data_row(sheet, key_columns: list[int]) -> int:
    for row_number in range(sheet.max_row, 1, -1):
        if any(sheet.cell(row_number, column).value not in (None, "") for column in key_columns):
            return row_number
    return 1


def _ensure_header(sheet, headers: dict[str, int], title: str) -> int:
    key = normalize(title)
    if key in headers:
        return headers[key]
    column = max(headers.values(), default=0) + 1
    source = sheet.cell(1, max(headers.values(), default=1))
    cell = sheet.cell(1, column, title)
    cell._style = copy(source._style)
    cell.font = copy(source.font)
    cell.fill = copy(source.fill)
    cell.alignment = copy(source.alignment)
    headers[key] = column
    return column


def append_received_document(extracted: dict, employee: dict, file_hash: str, enrichment_status: str, validation: dict | None = None) -> dict:
    """Acrescenta o documento à planilha configurada com gravação atômica."""
    target_path = configured_path("PIPELINE_ATESTADOS_PATH")
    with _workbook_write_lock, _spreadsheet_lease():
        workbook = load_workbook(target_path)
        temp_path = target_path.with_name(f".{target_path.stem}-{uuid.uuid4().hex}.tmp.xlsx")
        try:
            sheet = workbook["2026"] if "2026" in workbook.sheetnames else workbook.active
            headers = header_map(sheet)
            required = {"MATRICULA", "NOME", "TELEFONE", "EMAIL", "EMPRESA", "TIPODEDOCUMENTO"}
            if required - set(headers):
                raise RuntimeError("Cabecalhos obrigatorios ausentes na planilha de atestados")
            cpf_column = _ensure_header(sheet, headers, "CPF")
            hash_column = _ensure_header(sheet, headers, "ID TÉCNICO DO ARQUIVO")
            status_column = _ensure_header(sheet, headers, "STATUS DO CRUZAMENTO")
            validation_column = _ensure_header(sheet, headers, "STATUS DA VALIDACAO")
            sheet.column_dimensions[sheet.cell(1, hash_column).column_letter].hidden = True

            existing_row = None
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, hash_column).value == file_hash:
                    existing_row = row
                    break

            key_columns = [headers["NOME"], headers["MATRICULA"], cpf_column]
            previous_row = _last_data_row(sheet, key_columns)
            row_number = existing_row or previous_row + 1
            if row_number > EXCEL_MAX_ROWS:
                raise RuntimeError("A planilha atingiu o limite máximo de linhas do Excel.")
            for column in range(1, sheet.max_column + 1):
                destination = sheet.cell(row_number, column)
                destination.value = None
                if previous_row > 1:
                    source = sheet.cell(previous_row, column)
                    destination._style = copy(source._style)
                    destination.number_format = source.number_format
                    destination.alignment = copy(source.alignment)
                    destination.protection = copy(source.protection)

            values = {
                "MATRICULA": employee.get("matricula"),
                "NOME": extracted.get("nome"),
                "TELEFONE": employee.get("telefone"),
                "EMAIL": employee.get("email"),
                "EMPRESA": employee.get("empresa"),
                "TIPODEDOCUMENTO": extracted.get("tipo_documento"),
                "DATADERECEBIMENTO": date.today(),
                "DATADOATESTADO": extracted.get("data_atestado"),
                "CID": extracted.get("cid"),
                "QUANTIDADEDEDIAS": extracted.get("dias_afastamento"),
            }
            for key, new_value in values.items():
                if key in headers:
                    sheet.cell(row_number, headers[key], safe_excel_value(new_value))
            sheet.cell(row_number, cpf_column, safe_excel_value(extracted.get("cpf")))
            sheet.cell(row_number, hash_column, safe_excel_value(file_hash))
            sheet.cell(row_number, status_column, safe_excel_value(enrichment_status))
            validation = validation or {}
            sheet.cell(
                row_number,
                validation_column,
                safe_excel_value(validation.get("label", "PENDENTE")),
            )
            if validation.get("inss_ping"):
                fill = PatternFill("solid", fgColor="FFC7CE")
            elif validation.get("errors"):
                fill = PatternFill("solid", fgColor="FFE699")
            else:
                fill = PatternFill("solid", fgColor="C6EFCE")
            for column in range(1, sheet.max_column + 1):
                sheet.cell(row_number, column).fill = fill
            workbook.save(temp_path)
            workbook.close()
            os.replace(temp_path, target_path)
            return {"status": "ATUALIZADO" if existing_row else "REGISTRADO", "linha": row_number}
        finally:
            try:
                workbook.close()
            finally:
                temp_path.unlink(missing_ok=True)


def remove_received_document(file_hash: str) -> bool:
    """Remove da planilha a linha técnica correspondente ao registro excluído."""
    if not file_hash:
        return False
    target_path = configured_path("PIPELINE_ATESTADOS_PATH")
    with _workbook_write_lock, _spreadsheet_lease():
        workbook = load_workbook(target_path)
        temp_path = target_path.with_name(f".{target_path.stem}-{uuid.uuid4().hex}.tmp.xlsx")
        try:
            sheet = workbook["2026"] if "2026" in workbook.sheetnames else workbook.active
            headers = header_map(sheet)
            hash_column = headers.get("IDTECNICODOARQUIVO")
            if not hash_column:
                return False
            for row in range(2, sheet.max_row + 1):
                if sheet.cell(row, hash_column).value == file_hash:
                    sheet.delete_rows(row, 1)
                    workbook.save(temp_path)
                    workbook.close()
                    os.replace(temp_path, target_path)
                    return True
            return False
        finally:
            try:
                workbook.close()
            finally:
                temp_path.unlink(missing_ok=True)

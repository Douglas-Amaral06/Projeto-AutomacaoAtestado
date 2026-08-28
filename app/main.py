import hashlib
import io
import json
import os
import secrets
import shutil
import sqlite3
import threading
import time
import uuid
from contextlib import asynccontextmanager
from datetime import datetime
from datetime import timedelta
from pathlib import Path
from typing import Literal

from dotenv import load_dotenv
from fastapi import FastAPI, File, Form, HTTPException, Request, UploadFile
from fastapi.middleware.cors import CORSMiddleware
from fastapi.middleware.trustedhost import TrustedHostMiddleware
from fastapi.responses import FileResponse, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.staticfiles import StaticFiles
from fastapi.templating import Jinja2Templates
from openpyxl import Workbook
from PIL import Image, UnidentifiedImageError
from pydantic import BaseModel, ConfigDict, Field

from .database import BASE_DIR, UPLOAD_DIR, connect, initialize_database
from .gemini_service import QuotaExceededError
from .maintenance import BACKUP_DIR, apply_retention, create_backup, detect_orphan_files, prune_backups
from .processing import QueueItemBusyError, add_log, process_queue_item, resume_pending_once, understandable_error
from .rate_limit import check_daily_quota, check_rate_limit
from .safe_errors import format_safe_error
from .spreadsheet_pipeline import append_received_document, find_employee, remove_received_document, safe_excel_value
from .validation import document_type, normalize_cid, normalize_cpf, validation_summary
from .security import (SESSION_COOKIE, attempt_retry_after, create_session, current_user,
                       encrypt_totp, hash_password, hash_token, is_login_blocked, login_key, login_keys,
                       permissions_for, record_login, is_attempt_blocked, require_csrf, require_permission,
                       trusted_client_ip, utc_now, verify_login_password,
                       verify_service_token)

load_dotenv(BASE_DIR / ".env")
initialize_database()
templates = Jinja2Templates(directory=BASE_DIR / "app" / "templates")
_worker_stop = threading.Event()


def positive_env_int(name: str, default: int) -> int:
    try:
        value = int(os.getenv(name, str(default)))
    except ValueError:
        return default
    return value if value > 0 else default


def format_template_date(value, include_time: bool = False) -> str:
    """Formata datas do banco para exibição sem depender de JavaScript."""
    if value in (None, ""):
        return "—"

    parsed = value
    if isinstance(value, str):
        try:
            parsed = datetime.fromisoformat(value.replace("Z", "+00:00"))
        except ValueError:
            return value

    if not hasattr(parsed, "strftime"):
        return str(value)

    pattern = "%d/%m/%Y %H:%M" if include_time else "%d/%m/%Y"
    return parsed.strftime(pattern)


templates.env.filters["date_br"] = format_template_date
templates.env.filters["datetime_br"] = lambda value: format_template_date(value, include_time=True)


def queue_worker():
    while not _worker_stop.is_set():
        try: resume_pending_once()
        except Exception: pass
        _worker_stop.wait(20)


def maintenance_worker():
    while not _worker_stop.is_set():
        try:
            interval = max(1, int(os.getenv("BACKUP_INTERVAL_HOURS", "24"))) * 3600
            newest = max((item.stat().st_mtime for item in BACKUP_DIR.glob("*.zip")), default=0)
            if time.time() - newest >= interval:
                backup = create_backup()
                retention = apply_retention()
                removed = prune_backups()
                orphans = detect_orphan_files()
                add_log("info", "manutencao_concluida", "Backup automatico verificado e manutencao executada", {
                    "backup": backup.name, "retencao": retention, "backups_removidos": removed,
                    "arquivos_orfaos": len(orphans),
                })
        except Exception as error:
            try:
                safe_message, safe_details = format_safe_error(error)
                add_log("erro", "manutencao_falhou", safe_message, safe_details)
            except Exception: pass
        _worker_stop.wait(3600)


@asynccontextmanager
async def lifespan(_app):
    _worker_stop.clear()
    thread = threading.Thread(target=queue_worker, daemon=True, name="fila-atestados")
    maintenance = threading.Thread(target=maintenance_worker, daemon=True, name="manutencao-atestados")
    thread.start()
    maintenance.start()
    yield
    _worker_stop.set()


app = FastAPI(title="Recebimento Seguro de Atestados", docs_url=None, redoc_url=None, lifespan=lifespan)
app.add_middleware(TrustedHostMiddleware, allowed_hosts=os.getenv("ALLOWED_HOSTS", "127.0.0.1,localhost").split(","))
app.add_middleware(CORSMiddleware, allow_origins=[], allow_origin_regex=r"chrome-extension://.*", allow_methods=["POST","GET"], allow_headers=["Content-Type","X-API-Token"])
app.mount("/static", StaticFiles(directory=BASE_DIR / "app" / "static", html=False), name="static")
ALLOWED_TYPES = {"application/pdf", "image/jpeg", "image/png"}
MAX_UPLOAD_BYTES = 15 * 1024 * 1024
MAX_MULTIPART_BYTES = MAX_UPLOAD_BYTES + 1024 * 1024


@app.exception_handler(Exception)
async def unexpected_exception_handler(_request: Request, error: Exception):
    """Impede que exceções não tratadas exponham mensagens no FastAPI/Uvicorn."""
    safe_message, safe_details = format_safe_error(error)
    try:
        add_log("erro", "falha_interna", safe_message, safe_details)
    except Exception:
        pass
    return JSONResponse(
        status_code=500,
        content={"detail": {"codigo": "internal_error", "mensagem": safe_message}},
    )


def detected_mime(path: Path) -> str | None:
    head = path.read_bytes()[:16]
    if head.startswith(b"%PDF-"): return "application/pdf"
    if head.startswith(b"\xff\xd8\xff"): return "image/jpeg"
    if head.startswith(b"\x89PNG\r\n\x1a\n"): return "image/png"
    if head.startswith(b"RIFF") and head[8:12] == b"WEBP": return "image/webp"
    return None


def validate_document_structure(path: Path, mime_type: str) -> None:
    if path.stat().st_size == 0:
        raise HTTPException(400, "O arquivo está vazio.")
    if mime_type == "application/pdf":
        with path.open("rb") as source:
            source.seek(max(0, path.stat().st_size - 2048))
            if b"%%EOF" not in source.read():
                raise HTTPException(400, "O PDF está incompleto ou corrompido.")
        return
    try:
        with Image.open(path) as image:
            image.verify()
            if image.width <= 0 or image.height <= 0:
                raise ValueError("dimensões inválidas")
    except (UnidentifiedImageError, OSError, ValueError) as error:
        raise HTTPException(400, "A imagem está corrompida ou incompleta.") from error


def require_admin(user) -> None:
    if user["perfil"] != "admin": raise HTTPException(403, "Acesso exclusivo do administrador")


def parse_optional_boolean(value: str, field_name: str) -> bool | None:
    if value == "":
        return None
    if value == "true":
        return True
    if value == "false":
        return False
    raise HTTPException(422, f"Valor inválido para {field_name}.")


def review_context(request: Request, item, user, values: dict | None = None, validation: dict | None = None):
    record = dict(item)
    if values:
        record.update(values)
    if record.get("arquivo_hash"):
        with connect() as connection:
            record["possivel_repeticao"] = connection.execute(
                "SELECT COUNT(*) FROM atestados WHERE arquivo_hash=? AND id<>?",
                (record["arquivo_hash"], record.get("id") or 0),
            ).fetchone()[0] > 0
    validation = validation or validation_summary(record)
    return {
        "item": record, "user": user, "csrf": user["csrf_token"],
        "validation": validation,
    }


@app.middleware("http")
async def security_headers(request: Request, call_next):
    response = await call_next(request)
    response.headers.update({
        "X-Content-Type-Options": "nosniff", "X-Frame-Options": "SAMEORIGIN",
        "Referrer-Policy": "no-referrer", "Permissions-Policy": "camera=(), microphone=(), geolocation=()",
        "Content-Security-Policy": "default-src 'self'; script-src 'self'; style-src 'self' 'unsafe-inline'; img-src 'self' data:; connect-src 'self'; frame-src 'self'; frame-ancestors 'self'; form-action 'self'; base-uri 'none'; object-src 'none'",
        "Cache-Control": "no-store",
    })
    if os.getenv("COOKIE_SECURE", "false").lower() == "true":
        response.headers["Strict-Transport-Security"] = "max-age=31536000; includeSubDomains"
    return response


@app.middleware("http")
async def upload_limits(request: Request, call_next):
    """Autentica e limita uploads antes de o FastAPI analisar o multipart."""
    if request.method == "POST" and request.url.path == "/api/atestados":
        try:
            token_id = verify_service_token(request)
            declared_size = request.headers.get("content-length", "")
            if not declared_size.isdigit():
                raise HTTPException(411, "Content-Length é obrigatório para uploads.")
            incoming_bytes = int(declared_size)
            if incoming_bytes <= 0:
                raise HTTPException(400, "Upload vazio.")
            if incoming_bytes > MAX_MULTIPART_BYTES:
                raise HTTPException(413, "Requisição de upload acima do limite permitido.")
            check_rate_limit(
                str(token_id),
                limit=positive_env_int("UPLOAD_RATE_LIMIT_PER_HOUR", 30),
                window_seconds=3600,
            )
            check_daily_quota(
                str(token_id),
                incoming_bytes,
                max_bytes=positive_env_int("UPLOAD_DAILY_QUOTA_MB", 300) * 1024 * 1024,
            )
            request.state.upload_token_id = token_id
        except HTTPException as error:
            return JSONResponse(
                status_code=error.status_code,
                content={"detail": error.detail},
                headers=error.headers,
            )
    return await call_next(request)


@app.middleware("http")
async def trusted_proxy_guard(request: Request, call_next):
    if os.getenv("TRUST_CLOUDFLARE", "false").casefold() == "true":
        try:
            trusted_client_ip(request)
        except HTTPException as error:
            return JSONResponse(
                status_code=error.status_code,
                content={"detail": error.detail},
            )
    return await call_next(request)


def web_user(request: Request):
    try: return current_user(request)
    except HTTPException: return None


def secure_request(request: Request) -> bool:
    if os.getenv("COOKIE_SECURE", "false").lower() == "true" or request.url.scheme == "https":
        return True
    return (
        os.getenv("TRUST_CLOUDFLARE", "false").lower() == "true"
        and request.headers.get("cf-visitor", "").find('"scheme":"https"') >= 0
    )


@app.get("/login", response_class=HTMLResponse)
def login_page(request: Request, erro: str = ""):
    if web_user(request): return RedirectResponse("/", 303)
    return templates.TemplateResponse(request=request, name="login.html", context={"erro": erro})


@app.post("/login")
def login(request: Request, usuario: str = Form(..., min_length=1, max_length=100), senha: str = Form(..., min_length=1, max_length=256)):
    ip_key, account_key = login_keys(request, usuario)
    if is_attempt_blocked(ip_key, 10, 15) or is_login_blocked(account_key):
        retry = max(attempt_retry_after(ip_key, 15), attempt_retry_after(account_key, 15))
        add_log("aviso", "login_bloqueado", "Tentativas excessivas de login bloqueadas", {"aguarde_segundos": retry})
        return RedirectResponse("/login?erro=Bloqueado+temporariamente.+Tente+novamente+mais+tarde", 303)
    with connect() as connection:
        user = connection.execute("SELECT * FROM usuarios WHERE usuario=? AND ativo=1", (usuario.strip(),)).fetchone()
    valid = verify_login_password(user, senha)
    record_login(ip_key, valid)
    record_login(account_key, valid)
    if not valid: return RedirectResponse("/login?erro=Credenciais+invalidas", 303)
    token, _, expires = create_session(user["id"], request)
    with connect() as connection: connection.execute("UPDATE usuarios SET ultimo_login=? WHERE id=?", (utc_now().isoformat(), user["id"]))
    response = RedirectResponse("/", 303)
    response.set_cookie(SESSION_COOKIE, token, httponly=True, secure=secure_request(request), samesite="strict", max_age=8*3600, path="/")
    add_log("info", "login", f"Login realizado por usuario #{user['id']}")
    return response


@app.post("/logout")
def logout(request: Request, csrf_token: str = Form(...)):
    user = current_user(request); require_csrf(request, user, csrf_token)
    with connect() as connection: connection.execute("DELETE FROM sessoes WHERE id=?", (user["session_id"],))
    response = RedirectResponse("/login", 303); response.delete_cookie(SESSION_COOKIE); return response


@app.get("/usuarios", response_class=HTMLResponse)
def users_page(request: Request):
    user=web_user(request)
    if not user:return RedirectResponse("/login",303)
    require_admin(user)
    with connect() as connection: rows=connection.execute("SELECT id,usuario,nome,perfil,ativo,criado_em,ultimo_login FROM usuarios ORDER BY nome").fetchall()
    return templates.TemplateResponse(request=request,name="users.html",context={"usuarios":rows,"user":user,"csrf":user["csrf_token"],"provisioning_uri":None})


@app.post("/usuarios")
def create_user(request:Request,csrf_token:str=Form(..., max_length=100),usuario:str=Form(..., min_length=1, max_length=100),nome:str=Form(..., min_length=1, max_length=200),senha:str=Form(..., min_length=12, max_length=256),perfil:str=Form(..., pattern="^(admin|analista)$")):
    admin=current_user(request); require_csrf(request,admin,csrf_token); require_admin(admin)
    if perfil not in {"admin","analista"}:raise HTTPException(400,"Perfil invalido")
    secret=secrets.token_urlsafe(32)
    try:
        with connect() as connection: uid=connection.execute("INSERT INTO usuarios(usuario,nome,senha_hash,totp_secret_encrypted,perfil) VALUES(?,?,?,?,?)",(usuario.strip(),nome.strip(),hash_password(senha),encrypt_totp(secret),perfil)).lastrowid
    except Exception as error:
        raise HTTPException(400,"Usuario existente ou dados invalidos") from error
    add_log("info","usuario_criado",f"Usuario #{uid} criado pelo administrador #{admin['id']}")
    with connect() as connection: rows=connection.execute("SELECT id,usuario,nome,perfil,ativo,criado_em,ultimo_login FROM usuarios ORDER BY nome").fetchall()
    return templates.TemplateResponse(request=request,name="users.html",context={"usuarios":rows,"user":admin,"csrf":admin["csrf_token"],"provisioning_uri":None})


@app.post("/usuarios/{user_id}/alternar")
def toggle_user(user_id:int,request:Request,csrf_token:str=Form(...)):
    admin=current_user(request);require_csrf(request,admin,csrf_token);require_admin(admin)
    if user_id==admin["id"]:raise HTTPException(400,"Nao desative a propria conta")
    with connect() as connection:
        target=connection.execute("SELECT perfil,ativo FROM usuarios WHERE id=?",(user_id,)).fetchone()
        if not target:raise HTTPException(404,"Usuario inexistente")
        if target["perfil"]=="admin" and target["ativo"]:
            active_admins=connection.execute("SELECT COUNT(*) FROM usuarios WHERE perfil='admin' AND ativo=1").fetchone()[0]
            if active_admins<=1:raise HTTPException(400,"O sistema deve manter ao menos um administrador ativo")
        connection.execute("UPDATE usuarios SET ativo=CASE ativo WHEN 1 THEN 0 ELSE 1 END WHERE id=?",(user_id,))
        connection.execute("DELETE FROM sessoes WHERE usuario_id=?",(user_id,))
    add_log("info","usuario_alternado",f"Status do usuario #{user_id} alterado por #{admin['id']}");return RedirectResponse("/usuarios",303)


@app.get("/extensao", response_class=HTMLResponse)
def extension_pairing_page(request: Request):
    user=web_user(request)
    if not user:return RedirectResponse("/login",303)
    require_admin(user)
    with connect() as connection: tokens=connection.execute("SELECT id,nome,ativo,criado_em,ultimo_uso,expira_em FROM tokens_servico ORDER BY id DESC").fetchall()
    return templates.TemplateResponse(request=request,name="pairing.html",context={"user":user,"csrf":user["csrf_token"],"codigo":None,"tokens":tokens})


@app.post("/extensao/gerar-codigo", response_class=HTMLResponse)
def generate_pairing_code(request:Request,csrf_token:str=Form(...)):
    admin=current_user(request);require_csrf(request,admin,csrf_token);require_admin(admin)
    code=f"{secrets.randbelow(1_000_000):06d}"
    expires=utc_now()+timedelta(minutes=10)
    with connect() as connection:
        connection.execute("DELETE FROM codigos_pareamento WHERE criado_por=? AND usado_em IS NULL",(admin["id"],))
        connection.execute("INSERT INTO codigos_pareamento(codigo_hash,criado_por,expira_em) VALUES(?,?,?)",(hash_token(code),admin["id"],expires.isoformat()))
        tokens=connection.execute("SELECT id,nome,ativo,criado_em,ultimo_uso,expira_em FROM tokens_servico ORDER BY id DESC").fetchall()
    add_log("info","pareamento_criado",f"Codigo de pareamento criado pelo administrador #{admin['id']}")
    return templates.TemplateResponse(request=request,name="pairing.html",context={"user":admin,"csrf":admin["csrf_token"],"codigo":code,"tokens":tokens})


@app.post("/extensao/tokens/{token_id}/revogar")
def revoke_extension_token(token_id:int,request:Request,csrf_token:str=Form(...)):
    admin=current_user(request);require_csrf(request,admin,csrf_token);require_admin(admin)
    with connect() as connection:connection.execute("UPDATE tokens_servico SET ativo=0 WHERE id=?",(token_id,))
    add_log("aviso","token_revogado",f"Token de extensao #{token_id} revogado por #{admin['id']}");return RedirectResponse("/extensao",303)


@app.get("/", response_class=HTMLResponse)
def dashboard(request: Request, status: str = "", page: int = 1, per_page: int = 50):
    user = web_user(request)
    if not user: return RedirectResponse("/login", 303)
    page = max(1, page)
    per_page = min(100, max(10, per_page))
    offset = (page - 1) * per_page
    with connect() as connection:
        if status:
            total = connection.execute(
                "SELECT COUNT(*) FROM atestados WHERE status=?", (status,)
            ).fetchone()[0]
            rows = connection.execute(
                """SELECT a.*,u.nome revisor_nome FROM atestados a
                   LEFT JOIN usuarios u ON u.id=a.revisado_por
                   WHERE a.status=? ORDER BY a.id DESC LIMIT ? OFFSET ?""",
                (status, per_page, offset),
            ).fetchall()
        else:
            total = connection.execute("SELECT COUNT(*) FROM atestados").fetchone()[0]
            rows = connection.execute(
                """SELECT a.*,u.nome revisor_nome FROM atestados a
                   LEFT JOIN usuarios u ON u.id=a.revisado_por
                   ORDER BY a.id DESC LIMIT ? OFFSET ?""", (per_page, offset)
            ).fetchall()
        queue = connection.execute("SELECT status,COUNT(*) quantidade FROM fila_processamento GROUP BY status").fetchall()
        failed_items = connection.execute(
            """SELECT id,arquivo_original,status,tentativas,erro_amigavel,atualizado_em
               FROM fila_processamento WHERE status IN ('falhou','pausado_quota','aguardando_retentativa') ORDER BY atualizado_em DESC LIMIT 100"""
        ).fetchall()
        duplicate_hashes = {row[0] for row in connection.execute(
            "SELECT arquivo_hash FROM atestados WHERE arquivo_hash IS NOT NULL GROUP BY arquivo_hash HAVING COUNT(*)>1"
        ).fetchall()}
    atestados = []
    for row in rows:
        record = dict(row)
        record["possivel_repeticao"] = record.get("arquivo_hash") in duplicate_hashes
        atestados.append({**record, "validation": validation_summary(record)})
    failures = [
        {**dict(item), "mensagem": item["erro_amigavel"] or "A extração falhou. Reprocesse o documento ou consulte o suporte."}
        for item in failed_items
    ]
    return templates.TemplateResponse(request=request, name="dashboard.html", context={
        "atestados": atestados, "queue": queue, "failed_items": failures,
        "orphan_files": detect_orphan_files(), "permissions": permissions_for(user),
        "page": page, "per_page": per_page,
        "total_pages": max(1, (total + per_page - 1) // per_page), "status_filter": status,
        "user": user, "csrf": user["csrf_token"],
    })


@app.get("/atestados/{record_id}", response_class=HTMLResponse)
def review_page(record_id: int, request: Request):
    user = web_user(request)
    if not user: return RedirectResponse("/login",303)
    require_permission(user, "review")
    with connect() as connection: row=connection.execute("SELECT * FROM atestados WHERE id=?",(record_id,)).fetchone()
    if not row: raise HTTPException(404,"Registro inexistente")
    return templates.TemplateResponse(request=request,name="review.html",context=review_context(request, row, user))


@app.post("/atestados/{record_id}/revisar")
def review_document(
    record_id: int,
    request: Request,
    acao: str = Form(..., pattern="^(aprovar|rejeitar)$"),
    csrf_token: str = Form(..., max_length=100),
    versao_registro: str = Form(..., max_length=40),
    nome: str = Form("", max_length=200),
    cpf: str = Form("", max_length=20),
    cid: str = Form("", max_length=20),
    dias_afastamento: str = Form("", max_length=10),
    data_atestado: str = Form("", max_length=10),
    observacoes: str = Form("", max_length=4000),
    motivo_rejeicao: str = Form("", max_length=1000),
    matricula: str = Form("", max_length=80),
    telefone: str = Form("", max_length=80),
    email: str = Form("", max_length=254),
    empresa: str = Form("", max_length=200),
    tipo_documento: str = Form("", max_length=40),
    crm: str = Form("", max_length=30),
    crm_uf: str = Form("", max_length=2),
    assinado: str = Form("", max_length=5),
    carimbado: str = Form("", max_length=5),
):
    user=current_user(request); require_csrf(request,user,csrf_token)
    require_permission(user, "review")
    if acao not in {"aprovar","rejeitar"}: raise HTTPException(400,"Acao invalida")
    if acao=="rejeitar" and not motivo_rejeicao.strip(): raise HTTPException(400,"Informe o motivo")
    status="confirmado" if acao=="aprovar" else "rejeitado"
    with connect() as connection:
        before=connection.execute("SELECT * FROM atestados WHERE id=?",(record_id,)).fetchone()
    if not before: raise HTTPException(404,"Registro inexistente")
    try:
        employee, enrichment_status = find_employee(nome, cpf)
    except RuntimeError:
        employee, enrichment_status = None, "BASE_NAO_CONFIGURADA"
    employee = employee or {"matricula":matricula,"telefone":telefone,"email":email,"empresa":empresa}
    signed_value = parse_optional_boolean(assinado, "assinatura")
    stamped_value = parse_optional_boolean(carimbado, "carimbo")
    reviewed = {"nome":nome.strip() or None,"cpf":normalize_cpf(cpf),"cid":normalize_cid(cid),"dias_afastamento":dias_afastamento.strip() or None,"data_atestado":data_atestado.strip() or None,"tipo_documento":document_type(tipo_documento) or tipo_documento.strip() or None,"status_enriquecimento":enrichment_status,"crm":crm.strip() or None,"crm_uf":crm_uf.strip().upper() or None,"assinado":signed_value,"carimbado":stamped_value}
    validation = validation_summary(reviewed)
    if acao == "aprovar" and validation["errors"]:
        add_log("aviso", "aprovacao_bloqueada_validacao", f"Atestado #{record_id} exige correcao", {"erros": validation["errors"]})
        return templates.TemplateResponse(request=request, name="review.html", context=review_context(request, before, user, {**reviewed, **employee, "observacoes": observacoes, "motivo_rejeicao": motivo_rejeicao}, validation), status_code=422)
    days = int(dias_afastamento) if dias_afastamento.strip().isdigit() else None
    reviewed["dias_afastamento"] = days
    reservation = utc_now().isoformat()
    with connect() as connection:
        cursor = connection.execute(
            """UPDATE atestados SET revisado_em=? WHERE id=?
               AND COALESCE(revisado_em,criado_em)=?""",
            (reservation, record_id, versao_registro),
        )
    if cursor.rowcount != 1:
        raise HTTPException(409, "Este atestado foi alterado por outra pessoa. Reabra a tela antes de salvar.")
    try:
        if before["arquivo_hash"]:
            append_received_document(reviewed,employee,before["arquivo_hash"],enrichment_status,validation)
        with connect() as connection:
            cursor = connection.execute("""UPDATE atestados SET nome=?,cpf=?,cid=?,dias_afastamento=?,data_atestado=?,observacoes=?,status=?,motivo_rejeicao=?,revisado_por=?,revisado_em=?,matricula=?,telefone=?,email=?,empresa=?,tipo_documento=?,status_enriquecimento=?,crm=?,crm_uf=?,assinado=?,carimbado=? WHERE id=? AND revisado_em=?""",(nome.strip() or None,reviewed["cpf"],reviewed["cid"],days,data_atestado.strip() or None,observacoes.strip() or None,status,motivo_rejeicao.strip() or None,user["id"],reservation,employee.get("matricula") or None,employee.get("telefone") or None,employee.get("email") or None,employee.get("empresa") or None,reviewed["tipo_documento"],enrichment_status,reviewed["crm"],reviewed["crm_uf"],signed_value,stamped_value,record_id,reservation))
        if cursor.rowcount != 1:
            raise HTTPException(409, "A reserva de revisão expirou. Reabra a tela antes de salvar.")
    except Exception:
        with connect() as connection:
            connection.execute(
                "UPDATE atestados SET revisado_em=? WHERE id=? AND revisado_em=?",
                (before["revisado_em"], record_id, reservation),
            )
        raise
    changed_values = {**locals(), "assinado": signed_value, "carimbado": stamped_value, "crm_uf": reviewed["crm_uf"]}
    add_log("info","revisao",f"Atestado #{record_id} {status} por usuario #{user['id']}",{"campos_alterados":[k for k in ("nome","cpf","cid","dias_afastamento","data_atestado","observacoes","crm","crm_uf","assinado","carimbado") if str(before[k] if before[k] is not None else "")!=str(changed_values[k] if changed_values[k] is not None else "")]})
    return RedirectResponse("/",303)


@app.post("/atestados/{record_id}/excluir")
def delete_document(record_id:int, request:Request, csrf_token:str=Form(...)):
    user=current_user(request); require_csrf(request,user,csrf_token)
    require_permission(user, "delete")
    with connect() as connection:
        item=connection.execute("SELECT arquivo_salvo,arquivo_hash FROM atestados WHERE id=?",(record_id,)).fetchone()
        if not item:raise HTTPException(404,"Registro inexistente")
        connection.execute("DELETE FROM fila_processamento WHERE atestado_id=?",(record_id,))
        connection.execute("DELETE FROM atestados WHERE id=?",(record_id,))
        still_used=connection.execute("SELECT COUNT(*) FROM atestados WHERE arquivo_salvo=?",(item["arquivo_salvo"],)).fetchone()[0]
        same_content_remains=connection.execute("SELECT COUNT(*) FROM atestados WHERE arquivo_hash=?",(item["arquivo_hash"],)).fetchone()[0]
    if not still_used:
        (UPLOAD_DIR/item["arquivo_salvo"]).unlink(missing_ok=True)
    if not same_content_remains:
        try:
            remove_received_document(item["arquivo_hash"])
        except RuntimeError:
            add_log("aviso","planilha_nao_atualizada","Registro excluido do sistema, mas a planilha automatica nao estava configurada")
    add_log("aviso","atestado_excluido",f"Atestado #{record_id} e arquivo removidos pelo usuario #{user['id']}")
    return RedirectResponse("/",303)


@app.get("/atestados/{record_id}/arquivo")
def download_original(record_id:int, request:Request):
    user = web_user(request)
    if not user: return RedirectResponse("/login",303)
    require_permission(user, "review")
    with connect() as connection: row=connection.execute("SELECT arquivo_original,arquivo_salvo FROM atestados WHERE id=?",(record_id,)).fetchone()
    if not row: raise HTTPException(404,"Registro inexistente")
    return FileResponse(UPLOAD_DIR/row["arquivo_salvo"],filename=row["arquivo_original"],headers={"Content-Disposition":f"inline; filename=\"documento-{record_id}{Path(row['arquivo_original']).suffix}\""})


class LogEntry(BaseModel):
    model_config = ConfigDict(extra="forbid")
    nivel: Literal["info", "aviso", "erro"] = "info"
    evento: str = Field(min_length=1, max_length=80)
    mensagem: str = Field(min_length=1, max_length=1000)
    detalhes: dict | None = None


class PairingRequest(BaseModel):
    model_config = ConfigDict(extra="forbid")
    codigo: str = Field(pattern=r"^\d{6}$")
    nome: str = Field(default="Extensao Chrome", min_length=1, max_length=80)


@app.post("/api/parear")
def pair_extension(entry:PairingRequest,request:Request):
    key=login_key(request,"pareamento")
    if is_attempt_blocked(key,max_failures=3,window_minutes=30):
        retry=attempt_retry_after(key,30)
        add_log("aviso","pareamento_bloqueado","Origem bloqueada apos 3 tentativas invalidas de pareamento",{"aguarde_segundos":retry})
        raise HTTPException(429,{"codigo":"pareamento_bloqueado","mensagem":"Pareamento bloqueado apos 3 tentativas invalidas.","aguarde_segundos":retry})
    code=entry.codigo.strip()
    if len(code)!=6 or not code.isdigit():
        record_login(key,False);raise HTTPException(400,"Codigo invalido")
    with connect() as connection:
        row=connection.execute("SELECT * FROM codigos_pareamento WHERE codigo_hash=? AND usado_em IS NULL AND expira_em>?",(hash_token(code),utc_now().isoformat())).fetchone()
        if not row:
            record_login(key,False);raise HTTPException(401,"Codigo incorreto ou expirado")
        raw_token=secrets.token_urlsafe(48)
        token_id=connection.execute("INSERT INTO tokens_servico(nome,token_hash,criado_por,expira_em) VALUES(?,?,?,?)",(entry.nome[:80],hash_token(raw_token),row["criado_por"],(utc_now()+timedelta(days=90)).isoformat())).lastrowid
        connection.execute("UPDATE codigos_pareamento SET usado_em=? WHERE id=?",(utc_now().isoformat(),row["id"]))
    record_login(key,True);add_log("info","extensao_pareada",f"Extensao #{token_id} pareada")
    return {"token":raw_token,"token_id":token_id}


@app.post("/api/logs")
def create_log(entry:LogEntry, request:Request):
    verify_service_token(request); add_log(entry.nivel,entry.evento,entry.mensagem,entry.detalhes); return {"ok":True}


@app.get("/api/extensao/status")
def extension_status(request:Request):
    verify_service_token(request)
    return {"conectada":True}


@app.get("/logs",response_class=HTMLResponse)
def logs_page(request:Request):
    user=web_user(request)
    if not user:return RedirectResponse("/login",303)
    require_admin(user)
    with connect() as connection: rows=connection.execute("SELECT * FROM logs ORDER BY id DESC LIMIT 500").fetchall()
    return templates.TemplateResponse(request=request,name="logs.html",context={"logs":rows,"user":user,"csrf":user["csrf_token"]})


@app.post("/api/atestados")
def receive_document(
    request: Request,
    file: UploadFile = File(...),
    id_mensagem: str = Form("", max_length=200),
    id_conversa: str = Form("", max_length=200),
    whatsapp_remetente: str = Form("", max_length=80),
    data_recebimento: str = Form("", max_length=40),
    unidade: str = Form("", max_length=30),
):
    stored_path = None
    queue_id = None
    try:
        message_id = id_mensagem.strip()[:200] or None
        if message_id:
            with connect() as connection:
                previous = connection.execute(
                    "SELECT id,atestado_id FROM fila_processamento WHERE id_mensagem=?", (message_id,)
                ).fetchone()
            if previous:
                return {"id": previous["atestado_id"], "status": "duplicado", "fila_id": previous["id"]}
        if file.content_type not in ALLOWED_TYPES:
            raise HTTPException(400, "Envie um arquivo PDF, JPG ou PNG.")
        received_at = data_recebimento.strip()
        if received_at:
            try:
                datetime.fromisoformat(received_at.replace("Z", "+00:00"))
            except ValueError as error:
                raise HTTPException(422, "Data de recebimento inválida.") from error
        extensions = {"application/pdf": ".pdf", "image/jpeg": ".jpg", "image/png": ".png"}
        stored_name = f"{uuid.uuid4().hex}{extensions[file.content_type]}"
        stored_path = UPLOAD_DIR / stored_name
        written = 0
        digest_builder = hashlib.sha256()
        with stored_path.open("xb") as output:
            while chunk := file.file.read(1024 * 1024):
                written += len(chunk)
                if written > MAX_UPLOAD_BYTES:
                    raise HTTPException(413, "Arquivo acima de 15 MB.")
                digest_builder.update(chunk)
                output.write(chunk)
        gemini_input_limit = positive_env_int("GEMINI_MAX_DOCUMENT_MB", 8) * 1024 * 1024
        if written > gemini_input_limit:
            raise HTTPException(
                413,
                f"Documento acima do limite configurado para leitura ({gemini_input_limit // (1024 * 1024)} MB).",
            )
        actual_mime = detected_mime(stored_path)
        if actual_mime != file.content_type:
            raise HTTPException(400, "Conteúdo do arquivo não corresponde ao tipo informado.")
        validate_document_structure(stored_path, actual_mime)
        digest = digest_builder.hexdigest()
        with connect() as connection:
            possible_duplicate_count = connection.execute(
                "SELECT COUNT(*) FROM fila_processamento WHERE arquivo_hash=?", (digest,)
            ).fetchone()[0]
        with connect() as connection:
            cursor=connection.execute("""INSERT INTO fila_processamento(
                arquivo_hash,arquivo_original,arquivo_salvo,mime_type,status,
                id_mensagem,id_conversa,whatsapp_remetente,data_recebimento,unidade
            ) VALUES(?,?,?,?, 'aguardando_retentativa',?,?,?,?,?)""",(
                digest,Path(file.filename or "documento").name[:255],stored_name,file.content_type,
                message_id,id_conversa.strip()[:200] or None,
                whatsapp_remetente.strip()[:80] or None,received_at or None,
                unidade.strip().upper()[:30] or None,
            ))
            queue_id=cursor.lastrowid
    except sqlite3.IntegrityError:
        if stored_path:
            stored_path.unlink(missing_ok=True)
        with connect() as connection:
            previous = connection.execute(
                "SELECT id,atestado_id FROM fila_processamento WHERE id_mensagem=?", (message_id,)
            ).fetchone()
        if previous:
            return {"id": previous["atestado_id"], "status": "duplicado", "fila_id": previous["id"]}
        raise
    except Exception:
        if queue_id is None and stored_path:
            stored_path.unlink(missing_ok=True)
        raise
    finally:
        file.file.close()
    try:
        result = process_queue_item(queue_id)
        if possible_duplicate_count:
            result["possivel_repeticao"] = True
            result["aviso"] = "Possível documento repetido. O reenvio foi aceito para conferência."
        return result
    except QuotaExceededError as error: raise HTTPException(429,detail={"codigo":"gemini_quota_exceeded","mensagem":"Limite temporário do serviço de leitura atingido.","aguarde_segundos":error.retry_after}) from error
    except Exception as error: raise HTTPException(503,detail={"codigo":"enfileirado","mensagem":"Falha temporaria; arquivo preservado para nova tentativa.","fila_id":queue_id}) from error


@app.post("/fila/retomar")
def resume_queue(request:Request,csrf_token:str=Form(...)):
    user=current_user(request); require_csrf(request,user,csrf_token)
    require_admin(user)
    with connect() as connection: connection.execute("UPDATE fila_processamento SET status='aguardando_retentativa',disponivel_em=? WHERE status IN ('pausado_quota','falhou')",(utc_now().isoformat(),))
    add_log("info","fila_retomada",f"Fila retomada por usuario #{user['id']}"); return RedirectResponse("/",303)


@app.post("/fila/{queue_id}/reprocessar")
def reprocess_queue_item(queue_id: int, request: Request, csrf_token: str = Form(...)):
    user = current_user(request)
    require_csrf(request, user, csrf_token)
    require_permission(user, "reprocess")
    with connect() as connection:
        item = connection.execute(
            "SELECT status,atestado_id FROM fila_processamento WHERE id=?", (queue_id,)
        ).fetchone()
        if not item:
            raise HTTPException(404, "Item da fila inexistente")
        if item["atestado_id"] or item["status"] not in {"falhou", "pausado_quota", "aguardando_retentativa"}:
            raise HTTPException(409, "Somente extrações com falha podem ser reprocessadas")
        connection.execute(
            """UPDATE fila_processamento SET status='aguardando_retentativa',tentativas=0,
               disponivel_em=NULL,ultimo_erro=NULL,erro_amigavel=NULL,
               lock_token=NULL,lock_expires_em=NULL,atualizado_em=? WHERE id=?""",
            (utc_now().isoformat(), queue_id),
        )
    add_log("info", "reprocessamento_solicitado", f"Fila #{queue_id} reenviada por usuario #{user['id']}")
    try:
        process_queue_item(queue_id)
    except QueueItemBusyError as error:
        raise HTTPException(409, understandable_error(error)) from error
    except Exception:
        pass
    return RedirectResponse("/", 303)


@app.get("/relatorios",response_class=HTMLResponse)
def reports(request:Request):
    user=web_user(request)
    if not user:return RedirectResponse("/login",303)
    require_permission(user, "reports")
    with connect() as connection:
        summary=connection.execute("SELECT COUNT(*) total,SUM(CASE WHEN status='pendente' THEN 1 ELSE 0 END) pendentes,SUM(CASE WHEN status='confirmado' THEN 1 ELSE 0 END) confirmados,SUM(CASE WHEN status='rejeitado' THEN 1 ELSE 0 END) rejeitados,COALESCE(SUM(CASE WHEN status='confirmado' THEN dias_afastamento ELSE 0 END),0) dias FROM atestados").fetchone()
        monthly=connection.execute("SELECT substr(COALESCE(data_atestado,criado_em),1,7) mes,COUNT(*) quantidade,COALESCE(SUM(dias_afastamento),0) dias FROM atestados GROUP BY mes ORDER BY mes DESC LIMIT 12").fetchall()
        timing=connection.execute("SELECT ROUND(AVG((julianday(revisado_em)-julianday(criado_em))*24),2) horas FROM atestados WHERE revisado_em IS NOT NULL").fetchone()
    return templates.TemplateResponse(request=request,name="reports.html",context={"summary":summary,"monthly":monthly,"timing":timing,"user":user,"csrf":user["csrf_token"]})


@app.get("/exportar.xlsx")
def export_xlsx(request:Request):
    user = web_user(request)
    if not user:return RedirectResponse("/login",303)
    require_permission(user, "export")
    headers = ["Matricula","Nome","CPF","Telefone","E-mail","Empresa","Tipo de documento","CRM","UF CRM","CID","Assinado","Carimbado","Dias","Data","Status","Status do enriquecimento","Observacoes","Recebido em","Revisado em"]
    query = "SELECT matricula,nome,cpf,telefone,email,empresa,tipo_documento,crm,crm_uf,cid,assinado,carimbado,dias_afastamento,data_atestado,status,status_enriquecimento,observacoes,criado_em,revisado_em FROM atestados ORDER BY id"
    workbook = Workbook(write_only=True)
    sheet_number = 1
    sheet = workbook.create_sheet("Atestados")
    sheet.append(headers)
    rows_in_sheet = 1
    with connect() as connection:
        cursor = connection.execute(query)
        while batch := cursor.fetchmany(1000):
            for row in batch:
                if rows_in_sheet >= 1_048_576:
                    sheet_number += 1
                    sheet = workbook.create_sheet(f"Atestados {sheet_number}")
                    sheet.append(headers)
                    rows_in_sheet = 1
                sheet.append([safe_excel_value(value) for value in row])
                rows_in_sheet += 1
    output = io.BytesIO()
    try:
        workbook.save(output)
    finally:
        workbook.close()
    output.seek(0)
    return StreamingResponse(output,media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",headers={"Content-Disposition":"attachment; filename=atestados.xlsx"})
